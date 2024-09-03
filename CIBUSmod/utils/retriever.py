import os
import time
import weakref
import warnings
import pandas as pd
import numpy as np
import itertools
from openpyxl import load_workbook

from .misc import inv_dict

EMPTY = float("nan")

class ParameterRetriever:
    '''Class that retrieves parameters from an Excel file based on a flexible filtering approach.
    The Excel file needs to include a sheet named 'default' with columns named 'parameter' and 'value'
    including parameter names and corresponding values. Any filter columns should be named 'f_<name>'.
    Any other columns are not used by the ParameterRetriever. All rows without a value in the 'parameter'
    column are also ignored allowing e.g. headings in the parameter sheets.
    
    Parameters
    ----------
    name : str
        Name of the Excel file (excluding the .xlsx file extension) in the 'default data' folder storing
        parameter values.'''

    instances = weakref.WeakSet() # WeakSet of all ParameterRetriever instances
    data_path_default = None # Path to default data
    data_path_scenarios = None # Path to scenario data

    @classmethod
    def set_data_folder(cls, path, default_path=None, scenarios_path=None):
        # Set data path
        cls.data_path = _path_from_str(path)

        # Set default data path
        if default_path is not None:
            cls.data_path_default = _path_from_str(default_path)
        else:
            cls.data_path_default = os.path.join(cls.data_path, 'default')

        # Set scenario data path
        if scenarios_path is not None:
            cls.data_path_scenarios = _path_from_str(scenarios_path)
        else:
            cls.data_path_scenarios = os.path.join(cls.data_path, 'scenarios')

        # Read relation tables
        cls.update_relation_tables()

    @classmethod
    def get_rel(cls,from_col='',to_col=''):
        '''Returns a dict with values in 'from_col' as keys and 'to_col' as values'''

        try:
            rel = cls.relation_tables[from_col]
        except KeyError:
            sheet_dict = {
                c:t
                for t in cls.relation_tables
                for c in cls.relation_tables[t].columns
            }
            rel = cls.relation_tables[sheet_dict[from_col]]

        if len(set(rel[from_col])) < len(set(rel[to_col])):
            raise ValueError(f'Only one-to-one or many-to-one relations are allowed. Did you mean from_col={to_col}, to_col={from_col}?')
        
        return rel[[from_col,to_col]].set_index(from_col).to_dict()[to_col]

    @classmethod
    def update_relation_tables(cls):
        path = os.path.join(cls.data_path,'relation_tables.xlsx')
        try:
            cls.relation_tables = pd.read_excel(
                path,
                sheet_name=None, dtype=str
            )
        except FileNotFoundError:
            warnings.warn(f"Could not update relation tables. '{str(path)}' not found.")
    
    @classmethod
    def update_all_parameter_values(cls,scenario_workbooks=None,year=None,modules='all',pars='all'):

        if modules != 'all':
            if not isinstance(modules,list):
                modules = [modules]
        else:
            modules = [m.name for m in cls.instances]

        for pr in cls.instances:
            if pr.name in modules:
                if isinstance(pars, dict):
                    try:
                        pars_ = pars[pr.name]
                    except KeyError:
                        # If module not in pars dict all pars are updated
                        pars_ = 'all'
                else:
                    pars_ = pars

                pr.update_parameter_values(scenario_workbooks,year,pars_)
            else:
                # Update to default data
                pr.update_parameter_values()

    @classmethod
    def qry_stats(cls):
        import copy
        qry_log_all = {}
        set_time_all = 0
        for pr in cls.instances:
            set_time_all += pr.set_time
            for att in pr.qry_log:
                if '.'.join([pr.name,att]) in qry_log_all:
                    qry_log_all['.'.join([pr.name,att])]['lvls'].update(pr.qry_log[att]['lvls'])
                    qry_log_all['.'.join([pr.name,att])]['n'] += pr.qry_log[att]['n']
                    qry_log_all['.'.join([pr.name,att])]['time'] += pr.qry_log[att]['time']
                else:
                    qry_log_all['.'.join([pr.name,att])] = copy.deepcopy(pr.qry_log[att])
        
        df = pd.DataFrame.from_dict(qry_log_all, orient='index').sort_values('time', ascending=False)
        df['%-of-time'] = df['time'] / df['time'].sum() * 100
        qt = df['time'].sum()
        qm = round(np.floor(qt/60))
        qs = round(qt - qm*60)
        
        sm = round(np.floor(set_time_all/60))
        ss = round(set_time_all - sm*60)

        print(f"Total time setting filters: {sm} min {ss} sec")
        print(f"Total time querying parameters: {qm} min {qs} sec")
        print(df.iloc[:10,1:].round(1))
        return df

    def __init__(self, name, **kwargs):
        
        ParameterRetriever.instances.add(self)
        self.name = name

        # Read parameter dataframe with all columns as str except the 'value' column
        path = os.path.join(self.data_path_default, self.name + '.xlsx')
        self.data = _read_xl(path,'default')

        self.filters = {}
        self.qry_log = {}
        self.set_time = 0
        
        self.set(**kwargs)
        
    def __repr__(self):

        str1 = "\n".join([
            f"{key.replace('f_','')} : {str(value[0]) if len(value)==1 else 'List of ' + str(len(value))}"
            for key,value in  self.filters.items()
        ])
        unique_pars = self.data.index.get_level_values("parameter").unique()
        str2 = ", ".join(unique_pars[:min(len(unique_pars),10)]) + (" ..." if len(unique_pars) > 10 else "")

        return f"""
ParameterRetriever object with {len(self.data)} parameters in total.

Filters
-------
{str1}

Parameters
----------
{str2}
"""

    def __len__(self):
        return self.max_filter_length
        
    def set(self, **kwargs):
        '''Method to set filter values. Filters are supplied as keyword arguments and applies to columns in the Excel sheet
        named 'f_<key>'. If filters not present in the Excel sheet columns are supplied those are ignored and a warning is
        printed.
        
        Parameters
        ----------
        **kwargs : str or list of str
            If lists are supplied these need to have equal lengths (also to previously supplied filters) or length 1
            
        Returns
        -------
        Nothing. Updates the ParameterRetriever filters.'''
        t0 = time.process_time()
        # Set filter values supplied
        for key, value in kwargs.items():
            self.filters.update(
                {'f_'+key : value if not isinstance(value, str) else [value]}
            )
        
        # Get length of filter values
        filter_lens = [len(v) for v in self.filters.values()]
        distinct_filter_lens = set(filter_lens)
        self.max_filter_length = max(distinct_filter_lens) if len(self.filters) > 0 else 0

        if len(self.filters) == 0:
            self.selection = None
            return None

        if len(distinct_filter_lens) > 2 or (distinct_filter_lens == 2 and 1 not in distinct_filter_lens):
            # Raise error if any filter array is longer than 1 and shorter than the maximum filetr array length
            for key in kwargs.keys():
                self.filters.pop('f_'+key, None)
            raise ValueError('Lists of differing lengths supplied as filters')

        # Create selection index broadcasting length-1 filters
        self.selection = pd.MultiIndex.from_frame(
            pd.DataFrame(
                {
                    col: (labels * self.max_filter_length if len(labels) == 1 else labels)
                    for col, labels in self.filters.items()
                }
            )
        )

        self.set_time += time.process_time() - t0

    def clear(self):
        self.filters = {}
        self.set()

    def remove(self,keys):
        if not isinstance(keys, list):
            keys = [keys]
        for key in keys:
            try:
                self.filters.pop('f_'+key, None)
            except:
                pass
        
        self.set()

    def get(self, parameter, **kwargs):
        '''Method to get values of a parameter under the set filters.
        
        Parameters
        ----------
        parameter : str
            Name of the parameter to be retrieved
        **kwargs
            Keyword arguments to be passed on as filters to ParameterRetriever.set()
            
        Returns
        -------
        numpy.ndarray with length equal to the length of filter values. containing the parameter values for the defined filters '''
        
        self.set(**kwargs)

        t0 = time.process_time()
        if parameter in self.qry_log:
            self.qry_log[parameter]['lvls'].update(set(self.filters))
            self.qry_log[parameter]['n'] += 1
        else:
            self.qry_log.update({parameter : {'lvls' : set(self.filters), 'n' : 1, 'time' : 0}})

        result = _get_parameter_values(self.data, self.selection, parameter)

        # If NaNs are return print warning and some useful information
        if np.isnan(result).any():
            if self.selection is None:
                warnings.warn(f"NaN returned! No filters supplied and could not find a default value for '{parameter}' in {self.name}.xlsx.")
            elif np.isnan(result).all():
                str1 = f"NaNs returned! No value for '{parameter}' found in '{self.name}'.xlsx matching any of the supplied filters (n={len(self.selection)}): \n----------\n"
                n = min([len(self.selection),5])
                str2 = "\n----------\n".join([
                    "\n".join([
                        key + " = " + val
                        for key,val in zip(self.selection.names, sel)
                    ])
                    for sel in self.selection[0:n]
                ]) + ("\n ..." if n<len(self.selection) else "")
                warnings.warn(str1+str2)
            else:
                nan_sel = self.selection[np.isnan(result)]
                str1 = f"NaNs returned! No value for '{parameter}' found in '{self.name}'.xlsx for some of the supplied filters (n={len(nan_sel)}): \n----------\n"
                n = min([len(nan_sel),5])
                str2 = "\n----------\n".join([
                    "\n".join([
                        key + " = " + val
                        for key,val in zip(nan_sel.names, sel)
                    ])
                    for sel in nan_sel[0:n]
                ]) + ("\n ..." if n<len(nan_sel) else "")
                warnings.warn(str1+str2)

        # Make sure that length of result match length of selection
        # in cases where only filters not in data are used
        if self.selection is not None:
            sel_len = len(self.selection)
            if len(result) != sel_len:
                assert len(result) == 1
                result = np.repeat(result, sel_len)
        
        self.qry_log[parameter]['time'] += time.process_time() - t0
        return result

    def get_from_frame(self,parameter,df,**kwargs):
        '''Get parametervalues based on index and columns in supplied pandas.DataFrame'''

        if min(df.shape)<1:
            raise ValueError('DataFrame must have at least one row and one column')

        row_names = df.index.names
        col_names = df.columns.names

        filters = pd.merge(
            df.index.to_frame(index=False),
            df.columns.to_frame(index=False),
            how='cross'
        )
        # Make dict
        filters = {lvl:filters.loc[:,lvl].tolist() for lvl in row_names+col_names}

        result = pd.DataFrame(filters)

        try:
            result['value'] = self.get(parameter,**filters,**kwargs)
        except ValueError:
            # remove all filters with lengt > 1 and try again
            for key in list(self.filters):
                if len(self.filters[key])>1:
                    self.filters.pop(key, None)
            result['value'] = self.get(parameter,**filters,**kwargs)

        result = result.pivot(index=row_names,columns=col_names,values='value')

        return result.align(df, join='right')[0]

    def update_parameter_values(self,scenario_workbooks=None,year=None,pars='all'):
        '''Method to update parameter values in ParameterRetriever according to specified scenario workbooks and year.
        
        New parameter values are stored in a separate Excel file named '<scenario name>.xlsx' in a sheet with the
        same name as default parameter xlsx file. In the scenario sheet new values are defined in year columns with
        column names on the format 'y_<year>'. New parameter values can be defined in the Excel sheet for arbitrary
        years and the method linearly interpolates values between defined years.
        
        Values can be defined in ralative (i.e. a factor to multiply the default value with) or absolute terms by
        writing 'rel' or 'abs' respectively in a separate column named 'val_is'.
        
        Scenario values can be more general than default values (i.e. apply to several default values) but not
        more specific.
        
        Parameters
        ----------
        scenario : str or list of str
            Name(s) of scenarios to update parameter values according to. Scenarios later in the list
            will override earlier scenarios if the same parameter value is changed in multiple scnearios.
        year : str or int
            Year to update parameter values to
        pars : str or list of str
            Parameters to update. If pars='all', all available parameters will be updated.
            
        Returns
        -------
        Nothing. Updates ParameterRetriever parameter values.'''

        # Check pars input
        if pars != 'all':
            if not isinstance(pars,list):
                pars = [pars]
                
        # Get path to default data
        def_path = os.path.join(self.data_path_default, self.name + '.xlsx')

        if scenario_workbooks is None:
            self.data = _read_xl(def_path,'default')
            return

        year = int(year)
        if isinstance(scenario_workbooks,str):
            scenario_workbooks = [scenario_workbooks]

        # Read default parameter values
        data = _read_xl(def_path,'default')
        # Create pd.Series for updated parameter values
        updated_data = data.copy()

        # Go through all scenarios in consucutive orderd.
        # If the same parameter is updated in multiple scenarios
        # only the scenario that is latest in the list will have
        # an effect.
        for scn_wb in scenario_workbooks:

            # Get path to scenario data
            scn_path = os.path.join(self.data_path_scenarios, scn_wb + '.xlsx')

            if not os.path.isfile(scn_path):
                # If file does not exist print warning and continue
                warnings.warn(f"No scenario data workbook found on path {scn_path}")
                continue

            wb = load_workbook(scn_path, read_only=True)
            if not self.name in wb.sheetnames:
                # If sheet does not exist don't update anything
                wb._archive.close()
                continue
            wb._archive.close()

            # Read scenario parameter values
            scn_data_raw = _read_xl(scn_path,self.name)
            scn_data = scn_data_raw.copy()

            # Select parameters to update
            if pars != 'all':
                scn_data = scn_data[scn_data.index.get_level_values('parameter').isin(pars)]
                
            # If xlsx and sheet was found but contained no parameters, move to next scenario
            if len(scn_data) == 0:
                continue

            # Interpolate scenario parameter values for 'year'
            # Drop 'y_' prefix in columns and convert to int
            scn_data.columns = scn_data.columns.str.replace('y_','').astype(int)
            # Add missing years
            scn_data = scn_data.reindex(
                pd.Index(
                    range(
                        min(scn_data.columns.min(), year),
                        max(scn_data.columns.max(), year)+1
                    )
                ), axis=1
            )

            # Interpolate values for intermediate years and if neededpropagate first/last value of a
            # parameter backward/forward.
            scn_data = scn_data.interpolate(axis=1, limit_direction='forward')

            # Interpolation example:
            # ---------------------        ----------------------------------------------------------------------
            # par  2000  2005  2010  --->  par   2000  2001  2002  2003  2004  2005  2006  2007  2008  2009  2010
            # A    nan   1.0   1.5   --->  A     nan   nan   nan   nan   nan   1.0   1.1   1.2   1.3   1.4   1.5
            # B    1.0   1.5   nan   --->  B     1.0   1.1   1.2   1.3   1.4   1.5   1.5   1.5   1.5   1.5   1.5
            # C    1.0   0.5   1.0   --->  C     1.0   0.9   0.8   0.7   0.6   0.5   0.6   0.7   0.8   0.9   1.0
            # ---------------------        ----------------------------------------------------------------------

            # Select year
            scn_data = scn_data[year].rename('value')

            val_iss = scn_data.index.get_level_values('val_is').unique()

            # Create series to keep track of accessed rows in scenario data workbook
            scn_data_rows = pd.Series(
                range(len(scn_data)),
                index=scn_data.index
            )
            accessed_rows = []

            # Go through parameters defined in relative (rel) and absolute (abs) terms
            for val_is in [v for v in ['rel','abs'] if v in val_iss]:

                scn_data_ = scn_data.xs(val_is, level='val_is')
                scn_data_rows_ = scn_data_rows.xs(val_is, level='val_is')

                # Go through parameters in scenario and update values
                for parameter in scn_data_.index.get_level_values('parameter').unique():

                    # Create selection
                    try:
                        selection = self.data.xs(parameter, level='parameter', drop_level=False).index
                    except KeyError:
                        continue
                    scn_selection = selection.droplevel('parameter')
                    
                    # If no filter columns in scenarios sheet (i.e. values to update parameters apply universaly)
                    # then make sure that only one value is found and update accordingly
                    if len(scn_data_.index.names)==1:
                        assert np.isscalar(scn_data_.xs(parameter))
                        values = pd.Series(
                            scn_data_.xs(parameter),
                            index=selection
                        )
                        # Get accessed rows
                        accessed_rows.append(
                            np.atleast_1d(scn_data_rows_.xs(parameter))
                        )
                    else:
                        # Drop selection levels not in scenario filter columns
                        for lvl in (set(selection.names) - set(scn_data_.index.names)):
                            scn_selection = scn_selection.droplevel(lvl)

                        # Get scenario values
                        values = pd.Series(
                            _get_parameter_values(scn_data_, scn_selection, parameter),
                            index=selection
                        )
                        # Get accessed rows
                        accessed_rows.append(
                            np.unique(_get_parameter_values(scn_data_rows_, scn_selection, parameter))
                        )

                    # If in relative terms multiply with original value
                    if val_is=='rel':
                        values = data.loc[selection] * values

                    # Update values
                    updated_data.update(values)
                    
            # Check for data in scenario data workbook that was not accessed
            not_accessed_data = scn_data_raw.loc[scn_data_rows.index[~scn_data_rows.isin(np.concatenate(accessed_rows))], :]
            if len(not_accessed_data) > 0:
                warnings.warn(
f"""
Some scenario parameter value(s) did not match any default parameter value and thus did not have any effect. Check scenario data workbook!
Module: '{self.name}'
Scenario workbook: {scn_path}
----------------
{not_accessed_data.to_string()}
----------------
"""
                )

        self.data = updated_data

                
    def get_unique(self,filter,qry=None):
        '''Get unique values for specified filter(s) in parameter Excel sheet
        
        Parameters
        ----------
        filter : str or list of str
            filter(s) to get unique values for
        qry : str
            An optional query string filter the parameter sheet before finding
            unique filter values

        Returns
        -------
        numpy.array of unique values for filter if a str is supplied or
        pandas.DataFrame of unique combinations of filter (non-NaN) values with filter names as columns
        '''
        if qry is not None:
            df = self.data.reset_index().query(qry)
        else:
            df = self.data.reset_index()

        if isinstance(filter, list):
            res = df[['f_'+f for f in filter]].dropna(how='any').drop_duplicates()
            res.columns = filter
            return res
        else:
            res = df['f_'+filter].unique() 
            return res[~pd.isna(res)]
    
def _read_csv(path,parameter):
    df = pd.read_csv(path, dtype=str)

    cell1 = df.columns[0]
    if cell1.startswith('cols_as_filter'):
        # If 'cols_as_filter' keyword found in first cell
        # then stack data
        df = pd.read_csv(path, dtype=str, header=1)
        df.columns.name = cell1.replace('cols_as_filter: ','')
        f_cols = [c for c in df.columns if c.startswith("f_")]
        df = (
            df
            .set_index(f_cols)
            .stack()
            .rename(parameter)
            .to_frame()
            .reset_index()
        )

    if parameter not in df.columns:
        raise ValueError(f"'{parameter}' not found in '{path}'")
    f_cols = [c for c in df.columns if c.startswith("f_")]

    df = df.rename({parameter:'value'}, axis=1)
    df['parameter'] = parameter

    return df.loc[:,f_cols+['parameter','value']]
        
def _read_xl(path,sheet):
        idx = pd.IndexSlice
        # Read xl and set value columns to type float
        df = pd.read_excel(path, sheet_name=sheet, dtype=str)

        # Get parameters from any referenced csv-files
        if sheet=='default':

            # Get rows with reference to csv-files
            to_read_csv = df.value.str.contains('.csv', na=False)
            csvs_to_read = df.loc[to_read_csv,idx['parameter','value']]

            # Drop rows refering to csv-files from df
            df = df.loc[~to_read_csv]

            # Read csvs and append to df
            for parameter,csv_file in csvs_to_read.values:

                csv_path = os.path.join(os.path.dirname(path),csv_file)
                df_csv = _read_csv(csv_path,parameter)

                df = pd.concat([df,df_csv])

        # Only retain filter column(s), parameter column and value column(s)
        index_cols = [c for c in df.columns if c.startswith("f_") or c=="val_is"]
        value_cols = "value" if "value" in df.columns else [c for c in df.columns if c.startswith("y_")]

        df = (
            df.loc[lambda d: d["parameter"].notnull()]
            #.loc[lambda d: d["value"].notnull()]
            .set_index(index_cols + ["parameter"])[value_cols]
            .astype(float)
        )

        # Go thorough any filter columns with a ':', which indicates that these
        # are expresed on som aggregated level and should be translated
        # using relation tables.
        rel_filters = [f for f in df.index.names if ':' in f]
        if len(rel_filters)>0:
            for rf in rel_filters:
                # Get aggregated and target filter names
                f_from = rf.split(':')[0][2:]
                f_to = rf.split(':')[1]
                rel = inv_dict(ParameterRetriever.get_rel(f_to,f_from))

                # If target filter column does not exist create it
                if 'f_'+f_to not in df.index.names:
                    df_ = df.to_frame()
                    df_['f_'+f_to] = np.nan
                    df_ = df_.set_index('f_'+f_to, append=True)
                    df = df_['value']
                
                # Get rows
                df_w_rf = df.loc[~df.index.get_level_values(rf).isna(),:]
                df = df.loc[df.index.get_level_values(rf).isna(),:]

                # Make sure that the target filter column is empty
                if not df_w_rf.index.get_level_values('f_'+f_to).isna().all():
                    raise Exception('Target filter column has values')

                # Get index of relative filter and target filter
                rf_i = df.index.names.index(rf)
                tf_i = df.index.names.index('f_'+f_to)

                # Create new df where data values are propagaed across target filter values
                new_df = pd.Series(
                    {
                        tuple(
                            list(idx[:tf_i]) +
                            [f] +
                            list(idx[tf_i+1:])
                        ) : v
                        for idx, v in zip(df_w_rf.index, df_w_rf.values)
                        for f in rel[idx[rf_i]]
                    },
                    name = 'value'
                )
                new_df.index.names = df.index.names

                # Remove any rows with identical filters as rows in the big df
                sel = [
                    tuple(
                        list(idx[:rf_i]) +
                        [np.nan] +
                        list(idx[rf_i+1:])
                    ) not in df.index
                    for idx in
                    new_df.index
                ]
                new_df = new_df[sel]

                # Add new data to df and drop aggregated filter level
                df = pd.concat([df,new_df])
                df = df.droplevel(rf)

            # Make sure that 'parameter' is the last level in index.
            par_i = df.index.names.index('parameter')
            df = df.reorder_levels(df.index.names[:par_i]+df.index.names[par_i+1:]+[df.index.names[par_i]])

        # Raise error if duplicates found and print some usefull info
        if df.index.duplicated().any():
            dup = df.index[df.index.duplicated()].get_level_values("parameter")
            n = min(len(dup),5)
            str1 = f"One or more parameter(s) in '{path}' have identical filter columns (n={len(dup)}): "
            str2 = ", ".join(["'"+d+"'" for d in dup]) + (", ..." if n<len(dup) else "")
            raise ValueError(str1+str2)
                

        return df

def _path_from_str(str):
    path = ''
    for word in str.split('/'):
        path = os.path.join(path, word)
    return path

def _get_problem_data(data, index_cols, parameter):
    if not isinstance(data, pd.Series):
        raise ValueError(f"data should be a pandas.Series")
    if unknown_columns := set(index_cols) - set(data.index.names):
        raise ValueError(f"did not find index columns {unknown_columns} in data index")
    
    problem_data = data.xs(parameter, level="parameter")
    
    # Only chose rows where levels not filtered for are empty
    # If this is not possible return None
    null_cols = set(problem_data.index.names) - set(index_cols)
    for col in null_cols:
        if EMPTY in problem_data.index.get_level_values(col):
            problem_data = problem_data.xs(EMPTY, level=col)
        else:
            return None
        
    # Drop levels that are not filtered for
    for lvl in problem_data.index.names:
        if (problem_data.index.nlevels > 1) & problem_data.index.get_level_values(lvl).isna().all():
            problem_data = problem_data.droplevel(lvl)
        
    if len(problem_data.index.names) > 1:
        problem_data = problem_data.reorder_levels([c for c in index_cols if c in problem_data.index.names]) # NEW!!! [c for c in index_cols if c in problem_data.index.names]

    return problem_data


def _select_with_defaults(data, index, columns_to_take_default):
    partial_data = data
    partial_index = index

    for col in columns_to_take_default:
        try:
            partial_data = (
                partial_data.xs(EMPTY, level=col)
                if isinstance(partial_data.index, pd.MultiIndex)
                else partial_data[EMPTY]
            )
        except KeyError:
            # There is no default/empty key in this index column, which means
            # that ignoring the column is not possible.
            return partial_data[[]]
        try:
            partial_index = partial_index.droplevel(col)
        except ValueError:
            # Can't drop last level so don't
            partial_data = pd.Series(partial_data, index=partial_index.get_level_values(0))

    if len(partial_index.names) == 1:
        partial_index = partial_index.get_level_values(0) # pd.MultiIndex --> pd.Index

    partial_result = partial_data.reindex(partial_index).set_axis(index)
    
    return partial_result


def _select_allowing_any_k_defaults(data, index, k):
    # Generate all the (n choose k) ways to have k default columns
    results = pd.concat(
        [
            _select_with_defaults(data, index, default_cols)
            for default_cols in itertools.combinations(index.names, k)
        ],
        axis=1,
    )
    exactly_one_result = results.notnull().sum(axis=1) == 1
        
    # Get elements with exactly one result
    results = results[exactly_one_result].sum(axis=1)
    # Drop duplicate indexes to be able to merge back (these will have returned the same value anyway)
    results = results[~results.index.duplicated(keep='first')]
    
    return results

def _select_with_least_defaults(selection, problem_data):
    # Start with an empty result
    result = pd.Series(data=EMPTY, index=selection)

    # Fill in the blanks by successively using k = 0, ..., n default values,
    # where n is the number of index columns in the full problem..
    for k in range(len(selection.names) + 1):
        index_remainder = result[result.isnull()].index
        result = result.fillna(
            _select_allowing_any_k_defaults(problem_data, index_remainder, k)
        )

    return result

def _get_parameter_values(data, selection, parameter):

    if selection is not None:
        selection = selection.copy()
        # Drop filters not in data
        for lvl in set(selection.names)-set(data.droplevel('parameter').index.names):
            if (selection.nlevels > 1):
                selection = selection.droplevel(lvl)
            else:
                selection = None

    # If no filters supplied check if only one value can be returned
    # else return NaN
    if selection is None:
        result = data.xs(parameter, level="parameter")
        # Get rows with only NaNs in filter columns
        result = result.loc[result.index.to_frame().isna().all(axis=1)]
        if len(result) == 1:
            return result.values
        else:
            return np.array(EMPTY)

    # Get the data subset for the parameter in question,
    # and use the default for each dimension not specified in the selection.
    problem_data = _get_problem_data(data, selection.names, parameter)
    if problem_data is None:
        return np.array([EMPTY]*len(selection))

    # Drop filters not in data
    for lvl in set(selection.names)-set(problem_data.index.names):
        selection = selection.droplevel(lvl)
    
    assert problem_data.index.names == selection.names

    # Get unique selections to imporve performance
    selection_unique = selection.unique()

    result = _select_with_least_defaults(selection_unique, problem_data).reindex(selection)

    return result.values